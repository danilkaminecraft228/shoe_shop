"""
Скрипт импорта данных из Excel-файлов в базу данных shoe_shop.

Файлы должны находиться в папке import_data рядом с этим скриптом:
  - Tovar.xlsx
  - user_import.xlsx
  - Заказ_import.xlsx
  - Пункты выдачи_import.xlsx

Запуск:
    python seed_import.py

См. книгу "Python 3 и PyQt 6" гл. 24 (работа с базами данных).
"""

import os
import sys
from datetime import datetime, date

import openpyxl

# Добавляем родительскую папку в путь, чтобы импортировать database.py
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from database import create_connection, close_connection


# Путь к папке с Excel-файлами
IMPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "import_data")


def safe_date(value):
    """
    Преобразует значение из Excel в объект date.
    Если значение уже datetime или date — возвращает date.
    Если число — конвертирует из Excel-сериального номера.
    Если строка — пытается распарсить.
    При ошибке возвращает None.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial date: день 1 = 1900-01-01 (с учётом бага високосности)
        from datetime import timedelta
        base = datetime(1899, 12, 30)
        try:
            return (base + timedelta(days=int(value))).date()
        except (ValueError, OverflowError):
            return None
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def get_or_create(connection, table, column, value):
    """
    Получает ID записи по значению, или создаёт новую и возвращает её ID.
    """
    cursor = connection.cursor()
    cursor.execute(f"SELECT id FROM {table} WHERE {column} = %s", (value,))
    row = cursor.fetchone()
    if row:
        cursor.close()
        return row[0]
    cursor.execute(f"INSERT INTO {table} ({column}) VALUES (%s)", (value,))
    connection.commit()
    cursor.close()
    return cursor.lastrowid


def import_roles(connection):
    """Импорт ролей."""
    roles = ["Администратор", "Менеджер", "Авторизированный клиент"]
    cursor = connection.cursor()
    for role in roles:
        cursor.execute("INSERT IGNORE INTO roles (name) VALUES (%s)", (role,))
    connection.commit()
    cursor.close()
    print(f"Импортировано ролей: {len(roles)}")


def import_categories(connection, categories):
    """Импорт категорий товаров."""
    cursor = connection.cursor()
    count = 0
    for cat in categories:
        if cat:
            cursor.execute("INSERT IGNORE INTO categories (name) VALUES (%s)", (cat,))
            count += 1
    connection.commit()
    cursor.close()
    print(f"Импортировано категорий: {count}")


def import_suppliers(connection, suppliers):
    """Импорт поставщиков."""
    cursor = connection.cursor()
    count = 0
    for s in suppliers:
        if s:
            cursor.execute("INSERT IGNORE INTO suppliers (name) VALUES (%s)", (s,))
            count += 1
    connection.commit()
    cursor.close()
    print(f"Импортировано поставщиков: {count}")


def import_manufacturers(connection, manufacturers):
    """Импорт производителей."""
    cursor = connection.cursor()
    count = 0
    for m in manufacturers:
        if m:
            cursor.execute("INSERT IGNORE INTO manufacturers (name) VALUES (%s)", (m,))
            count += 1
    connection.commit()
    cursor.close()
    print(f"Импортировано производителей: {count}")


def import_products(connection, filepath):
    """
    Импорт товаров из Tovar.xlsx.
    Столбцы: Артикул, Наименование товара, Единица измерения, Цена,
             Поставщик, Производитель, Категория товара, Действующая скидка,
             Кол-во на складе, Описание товара, Фото
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    cursor = connection.cursor()
    count = 0
    errors = 0

    # Собираем все уникальные категории, поставщиков, производителей
    all_cats = set()
    all_supps = set()
    all_manufs = set()
    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        all_cats.add(row[6])
        all_supps.add(row[4])
        all_manufs.add(row[5])
        rows_data.append(row)

    # Сначала импортируем справочники
    import_categories(connection, all_cats)
    import_suppliers(connection, all_supps)
    import_manufacturers(connection, all_manufs)

    for row in rows_data:
        try:
            article = str(row[0]).strip()
            name = str(row[1]).strip() if row[1] else ""
            unit = str(row[2]).strip() if row[2] else "шт."
            price = float(row[3]) if row[3] is not None else 0
            supplier_name = str(row[4]).strip() if row[4] else ""
            manufacturer_name = str(row[5]).strip() if row[5] else ""
            category_name = str(row[6]).strip() if row[6] else ""
            discount = float(row[7]) if row[7] is not None else 0
            stock = int(row[8]) if row[8] is not None else 0
            description = str(row[9]).strip() if row[9] else ""
            image_path = str(row[10]).strip() if row[10] else None

            if not article:
                continue

            supplier_id = get_or_create(connection, "suppliers", "name", supplier_name)
            manufacturer_id = get_or_create(connection, "manufacturers", "name", manufacturer_name)
            category_id = get_or_create(connection, "categories", "name", category_name)

            cursor.execute("""
                INSERT INTO products
                (article, name, unit, price, supplier_id, manufacturer_id,
                 category_id, discount, stock_quantity, description, image_path)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    name = VALUES(name), unit = VALUES(unit), price = VALUES(price),
                    supplier_id = VALUES(supplier_id), manufacturer_id = VALUES(manufacturer_id),
                    category_id = VALUES(category_id), discount = VALUES(discount),
                    stock_quantity = VALUES(stock_quantity), description = VALUES(description),
                    image_path = VALUES(image_path)
            """, (article, name, unit, price, supplier_id, manufacturer_id,
                  category_id, discount, stock, description, image_path))
            count += 1
        except Exception as e:
            print(f"  Ошибка импорта товара (строка {count + 2}): {e}")
            errors += 1
            connection.rollback()
            continue

    connection.commit()
    cursor.close()
    wb.close()
    print(f"Импортировано товаров: {count}, ошибок: {errors}")


def import_users(connection, filepath):
    """
    Импорт пользователей из user_import.xlsx.
    Столбцы: Роль сотрудника, ФИО, Логин, Пароль
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    cursor = connection.cursor()
    count = 0
    errors = 0

    # Получаем словарь ролей
    role_map = {}
    cursor.execute("SELECT id, name FROM roles")
    for role_id, role_name in cursor.fetchall():
        role_map[role_name] = role_id

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        try:
            role_name = str(row[0]).strip() if row[0] else "Авторизированный клиент"
            full_name = str(row[1]).strip()
            login = str(row[2]).strip() if row[2] else ""
            password = str(row[3]).strip() if row[3] else ""

            if not login or not password:
                continue

            role_id = role_map.get(role_name, 3)  # 3 = Авторизированный клиент по умолч.

            cursor.execute("""
                INSERT INTO users (role_id, full_name, login, password)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    role_id = VALUES(role_id), full_name = VALUES(full_name),
                    password = VALUES(password)
            """, (role_id, full_name, login, password))
            count += 1
        except Exception as e:
            print(f"  Ошибка импорта пользователя: {e}")
            errors += 1
            connection.rollback()
            continue

    connection.commit()
    cursor.close()
    wb.close()
    print(f"Импортировано пользователей: {count}, ошибок: {errors}")


def import_pickup_points(connection, filepath):
    """
    Импорт пунктов выдачи из Пункты выдачи_import.xlsx.
    Одна колонка — адрес.
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    cursor = connection.cursor()
    count = 0

    for row in ws.iter_rows(min_row=1, values_only=True):
        address = str(row[0]).strip() if row[0] else ""
        if address:
            cursor.execute("INSERT INTO pickup_points (address) VALUES (%s)", (address,))
            count += 1

    connection.commit()
    cursor.close()
    wb.close()
    print(f"Импортировано пунктов выдачи: {count}")


def import_orders(connection, filepath):
    """
    Импорт заказов из Заказ_import.xlsx.

    Столбцы: Номер заказа, Артикул заказа, Дата заказа, Дата доставки,
             Адрес пункта выдачи, ФИО авторизированного клиента,
             Код для получения, Статус заказа

    Поле "Артикул заказа" вида: "А112Т4, 2, F635R4, 2"
    — чётные позиции: артикул, нечётные: количество.
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    cursor = connection.cursor()
    count = 0
    errors = 0
    skipped_dates = 0

    # Получаем маппинги
    user_map = {}
    cursor.execute("SELECT id, full_name FROM users")
    for uid, uname in cursor.fetchall():
        user_map[uname.strip().lower()] = uid

    status_map = {}
    cursor.execute("SELECT id, name FROM order_statuses")
    for sid, sname in cursor.fetchall():
        status_map[sname.strip().lower()] = sid

    address_map = {}
    cursor.execute("SELECT id, address FROM pickup_points")
    for pid, addr in cursor.fetchall():
        address_map[addr.strip().lower()] = pid

    product_map = {}
    cursor.execute("SELECT id, article FROM products")
    for pid, art in cursor.fetchall():
        product_map[art.strip().upper()] = pid

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        try:
            order_number = int(row[0]) if row[0] else None
            article_field = str(row[1]).strip() if row[1] else ""
            raw_order_date = row[2]
            raw_delivery_date = row[3]
            address_value = row[4]
            client_name = str(row[5]).strip() if row[5] else ""
            receive_code = str(row[6]).strip() if row[6] else ""
            status_name = str(row[7]).strip() if row[7] else "Новый"

            if not order_number:
                continue

            # Преобразуем даты
            order_date = safe_date(raw_order_date)
            delivery_date = safe_date(raw_delivery_date)

            # Валидация дат: год не должен быть ранее 2000 и позже 2100
            if order_date and (order_date.year < 2000 or order_date.year > 2100):
                print(f"  Предупреждение: некорректная дата заказа {order_date}, пропуск")
                order_date = None
                skipped_dates += 1
            if delivery_date and (delivery_date.year < 2000 or delivery_date.year > 2100):
                print(f"  Предупреждение: некорректная дата доставки {delivery_date}, пропуск")
                delivery_date = None
                skipped_dates += 1

            if order_date is None:
                # Если дата заказа некорректна, пропускаем запись
                print(f"  Пропуск заказа #{order_number}: некорректная дата заказа")
                errors += 1
                continue

            # Получаем ID пункта выдачи
            address_key = str(address_value).strip().lower() if address_value else ""
            pickup_id = address_map.get(address_key)
            if address_key and address_key not in address_map:
                # Пробуем по целочисленному значению (индекс из данных)
                if isinstance(address_value, (int, float)) and int(address_value) in address_map.values():
                    pickup_id = int(address_value)
                else:
                    # Создаём новый пункт
                    cursor.execute("INSERT INTO pickup_points (address) VALUES (%s)",
                                   (str(address_value).strip(),))
                    connection.commit()
                    pickup_id = cursor.lastrowid
                    address_map[address_key] = pickup_id

            if pickup_id is None:
                print(f"  Пропуск заказа #{order_number}: не указан пункт выдачи")
                errors += 1
                continue

            # Получаем ID клиента
            client_key = client_name.strip().lower()
            client_id = user_map.get(client_key)
            if client_id is None:
                # Если клиент не найден — создаём
                cursor.execute("""
                    INSERT INTO users (role_id, full_name, login, password)
                    VALUES (3, %s, %s, %s)
                """, (client_name, f"client_{order_number}@temp.com", "temp123"))
                connection.commit()
                client_id = cursor.lastrowid
                user_map[client_key] = client_id
                print(f"  Создан новый пользователь: {client_name}")

            # Получаем ID статуса
            status_key = status_name.strip().lower()
            status_id = status_map.get(status_key)
            if status_id is None:
                cursor.execute("INSERT INTO order_statuses (name) VALUES (%s)", (status_name,))
                connection.commit()
                status_id = cursor.lastrowid
                status_map[status_key] = status_id

            # Вставляем заказ
            cursor.execute("""
                INSERT INTO orders (id, order_date, delivery_date, pickup_point_id,
                                    client_id, receive_code, status_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    order_date = VALUES(order_date), delivery_date = VALUES(delivery_date),
                    pickup_point_id = VALUES(pickup_point_id), client_id = VALUES(client_id),
                    receive_code = VALUES(receive_code), status_id = VALUES(status_id)
            """, (order_number, order_date, delivery_date, pickup_id,
                  client_id, receive_code, status_id))

            # Разбираем состав заказа
            # Формат: "А112Т4, 2, F635R4, 2" — артикул, количество, артикул, количество...
            parts = [p.strip() for p in article_field.split(",") if p.strip()]
            i = 0
            while i < len(parts):
                article = parts[i].upper()
                qty = 1  # по умолчанию
                if i + 1 < len(parts):
                    try:
                        qty = int(float(parts[i + 1]))
                        i += 2  # использовали артикул и количество
                    except ValueError:
                        i += 1  # только артикул
                else:
                    i += 1

                product_id = product_map.get(article)
                if product_id is None:
                    print(f"  Предупреждение: товар с артикулом {article} не найден, пропуск")
                    continue

                cursor.execute("""
                    INSERT INTO order_items (order_id, product_id, quantity)
                    VALUES (%s, %s, %s)
                    ON DUPLICATE KEY UPDATE quantity = VALUES(quantity)
                """, (order_number, product_id, qty))

            count += 1
        except Exception as e:
            print(f"  Ошибка импорта заказа (строка {count + 2}): {e}")
            errors += 1
            connection.rollback()
            continue

    connection.commit()
    cursor.close()
    wb.close()
    print(f"Импортировано заказов: {count}, ошибок: {errors}, "
          f"пропущено по дате: {skipped_dates}")


def main():
    """Главная функция импорта."""
    print("=" * 60)
    print("Импорт данных в базу shoe_shop")
    print("=" * 60)

    if not os.path.exists(IMPORT_DIR):
        print(f"Папка с данными не найдена: {IMPORT_DIR}")
        print("Создайте папку import_data и поместите в неё Excel-файлы.")
        return

    connection = create_connection()
    if not connection:
        print("Не удалось подключиться к базе данных. Импорт прерван.")
        return

    try:
        # 1. Роли
        print("\n--- Импорт ролей ---")
        import_roles(connection)

        # 2. Пользователи (нужны для заказов)
        user_file = os.path.join(IMPORT_DIR, "user_import.xlsx")
        if os.path.exists(user_file):
            print("\n--- Импорт пользователей ---")
            import_users(connection, user_file)
        else:
            print(f"\nФайл не найден: {user_file}")

        # 3. Товары
        tovar_file = os.path.join(IMPORT_DIR, "Tovar.xlsx")
        if os.path.exists(tovar_file):
            print("\n--- Импорт товаров ---")
            import_products(connection, tovar_file)
        else:
            print(f"\nФайл не найден: {tovar_file}")

        # 4. Пункты выдачи
        points_file = os.path.join(IMPORT_DIR, "Пункты выдачи_import.xlsx")
        if os.path.exists(points_file):
            print("\n--- Импорт пунктов выдачи ---")
            import_pickup_points(connection, points_file)
        else:
            print(f"\nФайл не найден: {points_file}")

        # 5. Заказы
        orders_file = os.path.join(IMPORT_DIR, "Заказ_import.xlsx")
        if os.path.exists(orders_file):
            print("\n--- Импорт заказов ---")
            import_orders(connection, orders_file)
        else:
            print(f"\nФайл не найден: {orders_file}")

        print("\n" + "=" * 60)
        print("Импорт завершён!")
        print("=" * 60)

    except Exception as e:
        print(f"\nКритическая ошибка импорта: {e}")
        connection.rollback()

    finally:
        close_connection(connection)


if __name__ == "__main__":
    main()
